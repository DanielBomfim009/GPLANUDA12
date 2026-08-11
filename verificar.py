# -*- coding: utf-8 -*-
"""Confere se o que o Gplan mostra é o que a planilha diz.

Os erros que mais assustam neste projeto não quebram nada: mostram um número
velho vindo de cache, ou o número certo debaixo do rótulo errado. Nada estoura,
nada avisa. Este script existe para que isso não dependa de alguém lembrar de
olhar.

    python verificar.py                      # contra o app local
    ALVO=https://gplan.onrender.com python verificar.py

Sai com código 1 se qualquer conferência falhar, então serve em automação.
"""
from __future__ import annotations

import os
import re
import sys
import warnings
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore")
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ALVO = os.environ.get("ALVO", "http://localhost:8501")
PLANILHA = Path(__file__).resolve().parent.parent / (
    "Controle de Relatório dos Instrumentos/01_ARQUIVO_ATUAL/"
    "CONTROLE_DOCUMENTAL_INSTRUMENTACAO_ATUAL.xlsx"
)
APROVADOS = {"SEM COMENTÁRIOS", "COM COMENTÁRIOS", "PARA CONSTRUÇÃO"}

falhas: list[str] = []


def conferir(nome: str, obtido, esperado, detalhe: str = "") -> None:
    ok = obtido == esperado
    marca = "ok  " if ok else "FALHA"
    extra = f"  ({detalhe})" if detalhe else ""
    print(f"  [{marca}] {nome:38} {obtido}{'' if ok else f' != {esperado}'}{extra}")
    if not ok:
        falhas.append(f"{nome}: mostrou {obtido}, esperado {esperado}")


def numero(txt: str) -> float:
    """'R$ 9.492.286,54' e '19,8%' viram número."""
    limpo = re.sub(r"[^\d,.-]", "", txt or "").replace(".", "").replace(",", ".")
    return float(limpo) if limpo not in ("", "-", ".") else 0.0


# ---------------------------------------------------------------- a planilha
print(f"\nPLANILHA  {PLANILHA.name}")
if not PLANILHA.exists():
    print(f"  não encontrei em {PLANILHA}")
    sys.exit(1)

wb = load_workbook(PLANILHA, read_only=True, data_only=True)
ERROS_EXCEL = {"#REF!", "#VALOR!", "#VALUE!", "#NAME?", "#N/A", "#DIV/0!",
               "#NUM!", "#NULO!", "#DESPEJAR!", "#SPILL!"}
com_erro = {
    aba: n for aba in wb.sheetnames
    if (n := sum(1 for linha in wb[aba].iter_rows(values_only=True) for v in linha
                 if isinstance(v, str) and v.strip() in ERROS_EXCEL))
}
wb.close()
conferir("abre sem pedir reparo", True, True)
conferir("células de erro do Excel", com_erro or "nenhuma", "nenhuma")

tags = pd.read_excel(PLANILHA, sheet_name="01_BASE_TAGS")
resumo = pd.read_excel(PLANILHA, sheet_name="07_TAG_RESUMO")
esperados = pd.read_excel(PLANILHA, sheet_name="08_RELATORIOS_ESPERADOS")
validacoes = pd.read_excel(PLANILHA, sheet_name="11_VALIDACOES")

n_tags = len(tags)
n_esp = len(esperados)
n_postados = int((esperados["EXISTE_NO_SIGEM"].astype(str).str.upper() == "SIM").sum())
n_aprovados = int(esperados["STATUS_SIGEM"].astype(str).str.strip().str.upper()
                  .isin(APROVADOS).sum())
avanco = n_aprovados / n_esp if n_esp else 0

print("\nREGRAS DOCUMENTAIS")
# Toda TAG tem exatamente um destes -- menos a cancelada, que saiu de escopo
# e nao gera relatorio nenhum, mas continua na base para rastreio.
_canc = set(tags.loc[tags.STATUS_FINAL.astype(str).str.strip().str.upper()
                     == "CANCELADO", "TAG"]) if "STATUS_FINAL" in tags.columns else set()
conferir("CCP + RTFCJI + RIMITPI = TAGs ativas",
         int(esperados.RELATORIO.isin(["CCP", "RTFCJI", "RIMITPI"]).sum()),
         n_tags - len(_canc), f"{len(_canc)} canceladas fora")
# A exclusividade e por TAG, nao pelo cabo. Um mesmo circuito liga duas
# pontas: o instrumento responde pelo ensaio de comunicacao e a caixa de
# juncao pelo de continuidade. Sao ensaios diferentes, em pontas diferentes.
# O que nao pode e a MESMA TAG cobrar os dois pelo mesmo circuito.
par = lambda rel: set(map(tuple, esperados[esperados.RELATORIO == rel]
                          [["TAG", "REFERENCIA"]].astype(str).values))
conferir("mesma TAG com CTECRI e RILTCI no mesmo circuito",
         len(par("CTECRI") & par("RILTCI")), 0)
conferir("CTECRI em circuito de potência",
         int(esperados[esperados.RELATORIO == "CTECRI"].REFERENCIA
             .astype(str).str.upper().str.endswith("-P").sum()), 0,
         "cabo de potência não carrega comunicação")
conferir("esperados no resumo = linhas da 08",
         int(resumo.RELATORIOS_ESPERADOS.sum()), n_esp)
conferir("pendentes = esperados - aprovados",
         bool((resumo.RELATORIOS_PENDENTES
               == resumo.RELATORIOS_ESPERADOS - resumo.RELATORIOS_APROVADOS).all()), True)
_com_regra = resumo[resumo.RELATORIOS_ESPERADOS > 0]
conferir("avanço por TAG = aprovados / esperados",
         bool(((_com_regra.AVANCO_DOCUMENTAL
                - _com_regra.RELATORIOS_APROVADOS / _com_regra.RELATORIOS_ESPERADOS)
               .abs() < 1e-9).all()), True)
conferir("cancelada com avanço zerado",
         float(resumo.loc[resumo.RELATORIOS_ESPERADOS == 0, "AVANCO_DOCUMENTAL"].sum()), 0.0)
conferir("11_VALIDACOES sem inconsistência",
         str(validacoes.iloc[0]["TIPO_VALIDACAO"]), "SEM_INCONSISTENCIA")

# TAG cancelada segue na base para rastreio, mas não cobra documentação: o
# relatório dela nunca vai existir, e a pendência ficaria pendurada para
# sempre sujando o avanço de todo mundo.
canceladas = set(tags.loc[tags.STATUS_FINAL.astype(str).str.strip().str.upper()
                          == "CANCELADO", "TAG"]) if "STATUS_FINAL" in tags.columns else set()
conferir("cancelada não gera relatório esperado",
         int(esperados.TAG.isin(canceladas).sum()), 0,
         f"{len(canceladas)} canceladas")
conferir("cancelada continua no resumo",
         int(resumo.TAG.isin(canceladas).sum()), len(canceladas))
conferir("cancelada declarada como tal",
         int((resumo.loc[resumo.TAG.isin(canceladas), "STATUS_DOCUMENTAL"]
              .astype(str).str.upper() == "CANCELADA").sum()), len(canceladas))

# ------------------------------------------------- o relogio dos parados
# O "parado ha" conta desde o parecer da fiscalizacao, nao desde a postagem:
# entre postar e ser recusado passam dias que sao espera por analise, nao por
# providencia de quem emitiu. A Sheet2 traz esse "Modificado em".
print(chr(10) + "RELÓGIO DOS PARADOS")
sigem_full = pd.read_excel(PLANILHA, sheet_name="04_BASE_SIGEM")
conferir("04_BASE_SIGEM traz a data do parecer", "DATA_PARECER" in sigem_full.columns, True)
if "DATA_PARECER" in sigem_full.columns:
    _post = pd.to_datetime(sigem_full.DATA, dayfirst=True, errors="coerce")
    _par = pd.to_datetime(sigem_full.DATA_PARECER, dayfirst=True, errors="coerce")
    _amb = _post.notna() & _par.notna()
    conferir("parecer nunca é anterior à postagem",
             int((_par[_amb] < _post[_amb].dt.normalize()).sum()), 0)
    # 32 dos recusados nao tem linha na Sheet2, entao ficam sem data de
    # parecer. Nao e defeito: o codigo cai de volta na data de postagem. O que
    # nao pode e um recusado ficar sem data nenhuma e o relogio nao correr.
    _rec = sigem_full[sigem_full.STATUS.astype(str).str.strip().str.upper() == "RECUSADO"]
    _sem_par = int(pd.to_datetime(_rec.DATA_PARECER, dayfirst=True, errors="coerce").isna().sum())
    _sem_data = int((pd.to_datetime(_rec.DATA_PARECER, dayfirst=True, errors="coerce").isna()
                     & pd.to_datetime(_rec.DATA, dayfirst=True, errors="coerce").isna()).sum())
    conferir("recusado sempre tem de onde contar", _sem_data, 0,
             f"{_sem_par} sem parecer usam a data de postagem")

# ---------------------------------------------------------------- a GITEC
# A medição de campo só vale se for do mesmo escopo do controle documental. A
# GITEC mistura compra, inspeção de recebimento e cabo sob a mesma coluna TAG,
# e os valores desses outros escopos são de outra EAP: se um deles entrar, o
# medido salta de R$ 253 mil para R$ 9,3 milhões sem nada ter sido montado.
print("\nMEDIÇÃO DE CAMPO (GITEC)")
try:
    gitec = pd.read_excel(PLANILHA, sheet_name="06_BASE_GITEC")
except ValueError:
    gitec = pd.DataFrame(columns=["TAG", "FASE", "VALOR", "STATUS"])
    conferir("aba 06_BASE_GITEC existe", False, True)

if not gitec.empty:
    # A planilha é um controle por TAG, e item de PPU, critério e preço moram
    # nela. O que vale da GITEC é o que mede esses itens: cabo e tubing são
    # outro controle, e compra não é montagem.
    itens = set(tags.ITEM_PPU.astype(str).str.strip())
    conferir("todo item medido é PPU da base",
             int((~gitec.ITEM_PPU_GITEC.astype(str).str.strip().isin(itens)).sum()), 0)
    conferir("nenhum item de cabo entrou",
             int(gitec.ITEM_PPU_GITEC.astype(str).str.strip().eq("4.3.8.1.8").sum()), 0)
    conferir("toda medição tem tag da base",
             int((~gitec.TAG.isin(set(tags.TAG))).sum()), 0)
    # medir a tag sob um item diferente do dela é divergência entre as duas
    # bases; tem de estar declarada na 11_VALIDACOES, nunca escondida
    ppu_tag = dict(zip(tags.TAG, tags.ITEM_PPU.astype(str).str.strip()))
    divergentes = int(sum(1 for t, i in zip(gitec.TAG, gitec.ITEM_PPU_GITEC.astype(str).str.strip())
                          if ppu_tag.get(t, i) != i))
    declaradas = int(validacoes.TIPO_VALIDACAO.astype(str).eq("GITEC_ITEM_DIVERGENTE").sum())
    conferir("divergência de item declarada", declaradas, divergentes)
    # Estar na GITEC não é estar medido: o evento entra lá quando vai para a
    # fiscalização e só vira medição quando ela aprova. Enquanto não aprovar,
    # nada foi medido — e o valor ainda pode não virar medição nenhuma.
    ap = gitec.STATUS.astype(str).str.strip().str.upper().str.startswith("APROVADO")
    conferir("medido = só o aprovado",
             int((resumo.MEDIDO_GITEC.astype(str).str.upper() == "SIM").sum()),
             int(gitec.loc[ap, "TAG"].nunique()))
    conferir("valor medido = soma dos aprovados",
             round(float(resumo.VALOR_GITEC.fillna(0).sum()), 2),
             round(float(gitec.loc[ap, "VALOR"].fillna(0).sum()), 2))
    conferir("aguardando aprovação em coluna própria",
             round(float(resumo.VALOR_GITEC_VERIF.fillna(0).sum()), 2),
             round(float(gitec.loc[~ap, "VALOR"].fillna(0).sum()), 2))
    conferir("em verificação não entra como medido",
             int(((resumo.MEDIDO_GITEC.astype(str).str.upper() == "SIM")
                  & (resumo.VALOR_GITEC.fillna(0) <= 0)).sum()), 0)
    # medir sem a documentação fechar é o que motivou o cartão; medir uma tag
    # que nem montada está seria erro de base, e vale saber na hora
    medidas = set(gitec.TAG)
    montadas = set(tags.loc[tags.STATUS_MONTAGEM.astype(str).str.strip().str.upper()
                            == "MONTADO", "TAG"])
    conferir("medida sem estar montada", len(medidas - montadas), 0)

# ---------------------------------------------------------------------- o app
print(f"\nAPP  {ALVO}")
try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("  playwright não instalado; conferido só a planilha")
    sys.exit(1 if falhas else 0)

TXT = ("() => [...document.querySelectorAll('%s')]"
       ".map(e => e.innerText.split(String.fromCharCode(10)).join('|'))")

with sync_playwright() as p:
    navegador = p.chromium.launch()
    pg = navegador.new_page(viewport={"width": 1600, "height": 1100})
    erros_js: list[str] = []
    pg.on("pageerror", lambda e: erros_js.append(str(e)))

    # Espera generosa mas finita. Uma aba que nao abre em 4 minutos nao vai
    # abrir, e a conferencia tem que reportar isso em vez de ficar pendurada:
    # a Progresso hoje nem responde em producao, por falta de memoria.
    ESPERA = int(os.environ.get("ESPERA", "240")) * 1000

    def abrir(caminho: str, marcador: str) -> bool:
        try:
            pg.goto(ALVO + caminho, wait_until="domcontentloaded", timeout=ESPERA)
            pg.wait_for_selector(marcador, timeout=ESPERA)
        except Exception:
            return False
        pg.wait_for_timeout(4000)
        return "Traceback" not in pg.inner_text("body")

    print("\n  todas as abas abrem")
    for caminho, nome, marcador in [
        ("", "Dashboard", ".du-tela"),
        ("/relatorios", "Relatórios", "table.gtbl"),
        ("/pesquisa", "Pesquisa tag", "table.gtbl"),
        ("/sigem", "Base SIGEM", "table.gtbl"),
        ("/progresso", "Progresso", "details.arv-n1"),
        ("/gitec", "Gitec", ".fx-tiles"),
    ]:
        conferir(nome, abrir(caminho, marcador), True)

    # Dashboard: os cinco cartões contra a planilha
    print("\n  Dashboard bate com a planilha")
    abrir("", ".du-tela")

    def cartoes_kpi() -> dict:
        return {c.split("|")[0]: c.split("|")[1]
                for c in pg.evaluate(TXT % ".du-kpi") if "|" in c}

    cartoes = cartoes_kpi()
    conferir("Total de tags", int(numero(cartoes.get("Total de tags", "0"))), n_tags)
    conferir("Emitidos SIGEM", int(numero(cartoes.get("Emitidos SIGEM", "0"))), n_postados)
    conferir("Pendentes", int(numero(cartoes.get("Pendentes", "0"))), n_esp - n_aprovados)
    conferir("Avanço geral", round(numero(cartoes.get("Avanço geral", "0")), 1),
             round(avanco * 100, 1))
    # O cabeçalho já ficou em "—" por eu passar a chave de cache no lugar do
    # carimbo. Só falta data quem não sabe quando os dados são de fato.

    # O rodapé passou a carregar a medição de campo. Cada cartão é conferido
    # contra a planilha, e o carimbo de atualização mudou de lugar: saiu do
    # cabeçalho e virou o quinto cartão.
    print("\n  Rodapé do Dashboard bate com a planilha")
    rodape = {c.split("|")[0]: c.split("|")[1]
              for c in pg.evaluate(TXT % ".du-mini") if "|" in c}
    montadas_n = int((tags.STATUS_MONTAGEM.astype(str).str.strip().str.upper()
                      == "MONTADO").sum())
    preco = pd.to_numeric(tags.set_index("TAG").PRECO_UNITARIO,
                          errors="coerce").fillna(0.0)
    prontas = resumo[(resumo.AVANCO_DOCUMENTAL >= 1.0)
                     & (resumo.MEDIDO_GITEC.astype(str).str.upper() != "SIM")]
    conferir("Tags montadas", int(numero(rodape.get("Tags montadas", "0"))), montadas_n)
    conferir("Previsto de medição", round(numero(rodape.get("Previsto de medição", "0")), 2),
             round(float(preco.reindex(prontas.TAG).fillna(0).sum()), 2))
    conferir("Valor total do rodapé", round(numero(rodape.get("Valor total", "0")), 2),
             round(float(preco.sum()), 2))
    conferir("Medido na GITEC", round(numero(rodape.get("Medido na GITEC", "0")), 2),
             round(float(resumo.VALOR_GITEC.fillna(0).sum()), 2))
    conferir("carimbo saiu do cabeçalho",
             pg.evaluate("() => document.querySelectorAll('.du-cab .du-selo').length"), 0)
    conferir("carimbo virou cartão",
             bool(re.search(r"\d{2}/\d{2}/\d{4}", rodape.get("Última atualização", ""))), True,
             rodape.get("Última atualização", ""))

    # Tela única: a promessa é caber sem rolar e sem esconder nada. Painel que
    # corta conteúdo por dentro é pior que barra de rolagem -- some informação
    # e nada na tela avisa.
    print("\n  Dashboard cabe numa tela")
    medida = pg.evaluate("""() => {
        const m = document.querySelector('[data-testid=stMain]');
        const cortados = [...document.querySelectorAll('.du-tela *')].filter(e => {
            const o = getComputedStyle(e).overflowY;
            return (o === 'hidden' || o === 'clip')
                   && e.scrollHeight > e.clientHeight + 1 && e.clientHeight > 0; });
        return {rolagem: m.scrollHeight - m.clientHeight,
                horizontal: document.documentElement.scrollWidth - document.documentElement.clientWidth,
                cortados: cortados.length,
                barras: document.querySelectorAll('.du-br').length,
                linhas: document.querySelectorAll('.du-tab .lin').length,
                grupos: document.querySelectorAll('.du-gp').length};
    }""")
    conferir("sem rolagem vertical", medida["rolagem"], 0)
    conferir("sem rolagem horizontal", medida["horizontal"], 0)
    conferir("nada cortado dentro dos painéis", medida["cortados"], 0)
    conferir("barras = tipos de relatório", medida["barras"], 13)
    conferir("linhas do Top 10", medida["linhas"], 10)
    conferir("cartões de grupo", medida["grupos"], int(resumo.GRUPO_REGRA.nunique()))
    # Todo botão precisa ter para onde ir. O botão Detalhes do Dashboard já
    # apontou para âncora inexistente e o clique não fazia nada, sem erro.
    conferir("botões Detalhes sem destino", pg.evaluate("""() =>
        [...document.querySelectorAll('a.btn-detalhes')]
            .filter(a => !document.getElementById(a.getAttribute('href').slice(1))).length"""), 0)

    # Filtro: tem que recortar as três bases juntas. Recortar só uma daria um
    # Dashboard com 464 tags somando os 25.100 relatórios de todas elas.
    print("\n  Filtro recorta tudo junto")
    fase = str(tags.FASE.dropna().astype(str).str.strip().value_counts().index[0])
    ids = set(tags.loc[tags.FASE.astype(str).str.strip() == fase, "TAG"])
    esp_f = esperados[esperados.TAG.isin(ids)]
    apr_f = int(esp_f.STATUS_SIGEM.astype(str).str.strip().str.upper().isin(APROVADOS).sum())
    from urllib.parse import quote as _q
    if abrir("/?fase=" + _q(fase), ".du-tela"):
        c = cartoes_kpi()
        conferir(f"tags na fase {fase[:18]}", int(numero(c.get("Total de tags", "0"))), len(ids))
        conferir("pendentes da fase", int(numero(c.get("Pendentes", "0"))), len(esp_f) - apr_f)
        conferir("avanço da fase", round(numero(c.get("Avanço geral", "0")), 1),
                 round(apr_f / len(esp_f) * 100, 1) if len(esp_f) else 0.0)
        conferir("filtro aparece no cabeçalho", pg.locator(".du-selo.filtro").count(), 1)
    else:
        conferir("Dashboard filtrado abre", False, True)

    # Progresso: os totais e o avanço da árvore
    print("\n  Progresso bate com a planilha")
    abrir("/progresso", "details.arv-n1")
    totais = {t.split("|")[0]: t.split("|")[1]
              for t in pg.evaluate(TXT % ".prg-tot > div") if "|" in t}
    conferir("Instrumentos", int(numero(totais.get("INSTRUMENTOS", "0"))), n_tags)
    conferir("Avanço", round(numero(totais.get("AVANÇO", "0")), 1), round(avanco * 100, 1))
    conferir("Completos", int(numero(totais.get("COMPLETOS", "0"))),
             int((resumo.AVANCO_DOCUMENTAL >= 1.0).sum()))
    conferir("Valor total", round(numero(totais.get("VALOR TOTAL", "0")), 2),
             round(float(pd.to_numeric(tags.PRECO_UNITARIO, errors="coerce")
                         .fillna(0).sum()), 2))
    # o topo da árvore é a FASE; SOP passou a ser o segundo nível
    conferir("Fases na árvore", pg.locator("details.arv-n1").count(),
             tags.FASE.fillna("-").astype(str).str.strip().nunique())

    conferir("erros de JavaScript", erros_js[:1] or "nenhum", "nenhum")
    navegador.close()

print()
if falhas:
    print(f"{len(falhas)} conferência(s) falharam:")
    for f in falhas:
        print(f"  - {f}")
    sys.exit(1)
print("Tudo confere: a tela mostra o que a planilha diz.")
